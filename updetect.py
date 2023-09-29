import ccxt
import sys

from openpyxl import Workbook
import pprint


exchange = ccxt.binance()
markets = exchange.load_markets()

wb = Workbook(iso_dates=True)
sheet = wb.active

symbol = sys.argv[1]

data = exchange.fetch_ohlcv(symbol, "4h")
sheet.cell(row=1, column=1).value = symbol

y = 2
for stick in data:
    sheet.cell(row=y, column=1).value = stick[4]
    if y > 1:
        sheet.cell(
            row=y, column=2
        ).value = f"=({sheet.cell(row=y-1, column=1).coordinate}-{sheet.cell(row=y, column=1).coordinate})/{sheet.cell(row=y-1, column=1).coordinate}"
    y = y + 1


wb.save(f"{symbol}.xlsx")
